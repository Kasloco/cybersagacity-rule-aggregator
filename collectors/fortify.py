"""Collector for OpenText Fortify SCA static analysis rules.

Fortify SCA (Static Code Analyzer) is a commercial static analysis tool
by OpenText (formerly Micro Focus, HP). The rule set is proprietary and
not publicly available in a Git repository. Fortify rules are organized
into "rulepacks" and cover security vulnerabilities, code quality, and
configuration issues across many languages.

Public documentation is available at:
  - https://www.microfocus.com/documentation/fortify/
  - https://www.opentext.com/products/fortify-static-code-analyzer

Fortify can export findings in several formats:
  1. FPR (Fortify Project Results) — XML-based, contains audit findings
     with rule IDs, categories, and CWE mappings.
  2. SARIF — Fortify can export SARIF results which include rule metadata.
  3. XLSX/CSV — Spreadsheet exports with checker categories and descriptions.
  4. Fortify rulepack XML — Internal rulepack format with rule definitions.

This collector uses source_type='file' — it does NOT clone a repository.
Instead, it imports Fortify rules from local/exported files in these
formats:

To use this collector:
  a. Export Fortify rules/findings in SARIF, FPR (XML), CSV, or XLSX format.
  b. Place the files in the directory specified by the
     FORTIFY_RULES_DIR environment variable (default:
     /tmp/cybersagacity-rules/fortify/).
  c. Run the collector — it will parse all supported files in that directory.

The collector can be extended with a proper importer when OpenText
publishes a public rules repository.
"""

import os
import csv
import json
import logging
import re
import xml.etree.ElementTree as ET

from .base import BaseCollector

logger = logging.getLogger(__name__)

# Fortify impact → normalized severity
SEVERITY_MAP = {
    "critical": "critical",
    "high": "high",
    "medium": "medium",
    "moderate": "medium",
    "low": "low",
    "info": "info",
    "informational": "info",
}

# Fortify category → normalized category
CATEGORY_MAP = {
    "buffer overflow": "memory-safety",
    "buffer overflow (cwe-120)": "memory-safety",
    "code injection": "injection",
    "command injection": "injection",
    "cross-site request forgery": "csrf",
    "cross-site scripting": "xss",
    "denial of service": "dos",
    "file disclosure": "information-disclosure",
    "insecure randomness": "crypto",
    "insecure temporary file": "filesystem",
    "insecure transport": "crypto",
    "json injection": "injection",
    "ldap injection": "injection",
    "log forging": "injection",
    "missing authorization": "auth",
    "missing encryption": "crypto",
    "missing input validation": "input-validation",
    "missing xml validation": "input-validation",
    "null dereference": "correctness",
    "os command injection": "injection",
    "password management": "secrets",
    "path manipulation": "path-traversal",
    "poor error handling": "error-handling",
    "poor logging": "logging",
    "privacy violation": "privacy",
    "race condition": "concurrency",
    "resource injection": "injection",
    "security misconfiguration": "misconfiguration",
    "sql injection": "sqli",
    "xss": "xss",
    "xml external entity": "xxe",
    "xpath injection": "injection",
}


class FortifyCollector(BaseCollector):
    name = "fortify"
    display_name = "Fortify SCA (OpenText)"
    source_type = "file"
    source_url = (
        "https://www.microfocus.com/documentation/fortify/"
    )
    description = (
        "OpenText Fortify Static Code Analyzer is a commercial static "
        "analysis tool that detects security vulnerabilities across many "
        "languages. Rules are organized into rulepacks covering injection, "
        "XSS, crypto, path traversal, configuration, and more. Each finding "
        "maps to CWEs and OWASP categories. This collector imports rules from "
        "exported SARIF, FPR (XML), CSV, or XLSX files."
    )
    logo_url = "https://www.microfocus.com/etc/clientlibs/microfocus/clientlibs/base/img/favicon.ico"

    def collect_rules(self):
        """Import Fortify rules from local file exports.

        This collector reads from the directory specified by the
        FORTIFY_RULES_DIR environment variable (or the default clone_dir).
        Supported file formats: .sarif, .fpr, .xml, .csv, .xlsx

        See the module docstring for instructions on exporting Fortify rules.
        """
        rules_dir = os.environ.get(
            "FORTIFY_RULES_DIR",
            self.clone_dir,
        )

        if not os.path.isdir(rules_dir):
            logger.info(
                f"[fortify] Rules directory {rules_dir} not found. "
                f"Export Fortify rules (SARIF/FPR/CSV/XLSX) and place them there. "
                f"Set FORTIFY_RULES_DIR to override."
            )
            return

        count = 0
        for fname in sorted(os.listdir(rules_dir)):
            fpath = os.path.join(rules_dir, fname)
            if not os.path.isfile(fpath):
                continue

            if fname.endswith(".sarif") or (
                fname.endswith(".json") and "fortify" in fname.lower()
            ):
                count += self._parse_sarif(fpath)
            elif fname.endswith((".fpr", ".xml")):
                count += self._parse_fpr_xml(fpath)
            elif fname.endswith(".csv"):
                count += self._parse_csv(fpath)
            elif fname.endswith(".xlsx"):
                count += self._parse_xlsx(fpath)

        logger.info(f"[fortify] Imported {count} rules from {rules_dir}")

    def _parse_sarif(self, fpath):
        """Parse a SARIF file for Fortify rule definitions."""
        count = 0

        try:
            with open(fpath, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError):
            logger.warning(f"[fortify] Failed to parse SARIF: {fpath}")
            return 0

        for run in data.get("runs", []):
            driver = run.get("tool", {}).get("driver", {})
            rules = driver.get("rules", [])

            for rule in rules:
                rule_id = rule.get("id", "")
                if not rule_id:
                    continue

                if not rule_id.startswith("fortify:"):
                    full_rule_id = f"fortify:{rule_id}"
                else:
                    full_rule_id = rule_id

                properties = rule.get("properties", {})
                full_desc = rule.get("fullDescription", {}).get("text", "")
                short_desc = rule.get("shortDescription", {}).get("text", "")
                name = rule.get("name", rule_id)

                # Severity
                level = rule.get("defaultConfiguration", {}).get("level", "warning")
                severity_raw = properties.get("severity", properties.get("impact", level))
                severity = SEVERITY_MAP.get(str(severity_raw).lower(), "medium")

                # CWE
                cwe_ids = []
                cwe_raw = properties.get("cwe", properties.get("cweId", ""))
                if cwe_raw:
                    if isinstance(cwe_raw, list):
                        cwe_ids = [str(c) for c in cwe_raw]
                    else:
                        cwe_nums = re.findall(r"(\d+)", str(cwe_raw))
                        cwe_ids = [f"CWE-{n}" for n in cwe_nums]

                # Category
                fortify_category = properties.get("category", "")
                category = CATEGORY_MAP.get(
                    fortify_category.lower().strip(), "security"
                )

                # OWASP
                owasp_ids = []
                owasp_raw = properties.get("owasp", "")
                if owasp_raw:
                    if isinstance(owasp_raw, list):
                        owasp_ids = [str(o) for o in owasp_raw]
                    else:
                        owasp_ids = [str(owasp_raw)]

                tags = ["fortify", "sast", category]
                if fortify_category:
                    tags.append(fortify_category.lower().replace(" ", "-"))

                self.upsert(
                    rule_id=full_rule_id,
                    title=f"Fortify {name}: {short_desc}"[:500],
                    description=full_desc or short_desc,
                    severity=severity,
                    category=category,
                    language=properties.get("language", ""),
                    cwe_ids=cwe_ids,
                    owasp_ids=owasp_ids,
                    tags=tags,
                    source_file=fpath,
                    rule_content=json.dumps(rule, indent=2, default=str)[:50000],
                    rule_format="sarif",
                    metadata={
                        "checker_id": rule_id,
                        "checker_name": name,
                        "fortify_category": fortify_category,
                        "impact": properties.get("impact", ""),
                        "accuracy": properties.get("accuracy", ""),
                        "source": "sarif-import",
                    },
                )
                count += 1

        return count

    def _parse_fpr_xml(self, fpath):
        """Parse a Fortify FPR (XML) or rulepack XML file for rule definitions.

        FPR files contain an Audit section with issue instances, each
        referencing a rule by its InstanceID/CategoryID. Rulepack XML
        files contain RuleDefinition elements with metadata.
        """
        count = 0

        try:
            tree = ET.parse(fpath)
        except ET.ParseError:
            logger.warning(f"[fortify] Failed to parse XML/FPR: {fpath}")
            return 0

        root = tree.getroot()

        # FPR structure: <FPR> → <Description> → <MetaInfo> with ClassInfo
        # Rulepack structure: <RulePack> → <RuleDefinition> elements

        # Try rulepack format first
        ns = ""
        # Detect namespace if present
        if root.tag.startswith("{"):
            ns = root.tag.split("}")[0] + "}"

        # Look for RuleDefinition elements (rulepack format)
        rule_defs = root.findall(
            f".//{ns}RuleDefinition"
        ) or root.findall(".//RuleDefinition")

        if rule_defs:
            for rd in rule_defs:
                count += self._parse_rule_definition(rd, fpath)
        else:
            # Try FPR audit format — extract unique rule categories from findings
            count += self._parse_fpr_audit(root, ns, fpath)

        tree = None  # release
        return count

    def _parse_rule_definition(self, rd, fpath):
        """Parse a single RuleDefinition element from a rulepack XML."""
        ns = ""
        if rd.tag.startswith("{"):
            ns = rd.tag.split("}")[0] + "}"

        rule_id = (
            rd.get("id")
            or rd.get("ruleId")
            or self._xml_text(rd, f"{ns}RuleID")
            or self._xml_text(rd, "RuleID")
            or self._xml_text(rd, f"{ns}Id")
            or ""
        )
        if not rule_id:
            return 0

        full_rule_id = f"fortify:{rule_id}"

        # Extract metadata from child elements
        name = (
            self._xml_text(rd, f"{ns}Name")
            or self._xml_text(rd, "Name")
            or self._xml_text(rd, f"{ns}Title")
            or self._xml_text(rd, "Title")
            or rule_id
        )

        description = (
            self._xml_text(rd, f"{ns}Description")
            or self._xml_text(rd, "Description")
            or self._xml_text(rd, f"{ns}DescriptionBrief")
            or self._xml_text(rd, "DescriptionBrief")
            or ""
        )

        # Severity/Impact
        severity_raw = (
            self._xml_text(rd, f"{ns}Impact")
            or self._xml_text(rd, "Impact")
            or self._xml_text(rd, f"{ns}Severity")
            or self._xml_text(rd, "Severity")
            or "medium"
        )
        severity = SEVERITY_MAP.get(str(severity_raw).lower(), "medium")

        # Category
        fortify_category = (
            self._xml_text(rd, f"{ns}Category")
            or self._xml_text(rd, "Category")
            or self._xml_text(rd, f"{ns}Subcategory")
            or self._xml_text(rd, "Subcategory")
            or ""
        )
        category = CATEGORY_MAP.get(
            fortify_category.lower().strip(), "security"
        )

        # CWE
        cwe_ids = []
        cwe_raw = (
            self._xml_text(rd, f"{ns}CWE")
            or self._xml_text(rd, "CWE")
            or self._xml_text(rd, f"{ns}CWEId")
            or self._xml_text(rd, "CWEId")
            or ""
        )
        if cwe_raw:
            cwe_nums = re.findall(r"(\d+)", str(cwe_raw))
            cwe_ids = [f"CWE-{n}" for n in cwe_nums]

        language = (
            self._xml_text(rd, f"{ns}Language")
            or self._xml_text(rd, "Language")
            or ""
        )

        tags = ["fortify", "sast", category]
        if fortify_category:
            tags.append(fortify_category.lower().replace(" ", "-"))

        self.upsert(
            rule_id=full_rule_id,
            title=f"Fortify {name}"[:500],
            description=description[:2000],
            severity=severity,
            category=category,
            language=language,
            cwe_ids=cwe_ids,
            tags=tags,
            source_file=fpath,
            rule_content=ET.tostring(rd, encoding="unicode")[:50000] if rd else "",
            rule_format="xml",
            metadata={
                "checker_id": rule_id,
                "checker_name": name,
                "fortify_category": fortify_category,
                "impact": severity_raw,
                "source": "rulepack-xml-import",
            },
        )
        return 1

    def _parse_fpr_audit(self, root, ns, fpath):
        """Parse an FPR audit file, extracting unique rules from findings."""
        count = 0
        seen_rules = set()

        # FPR findings are in <Issue> or <IssueInstance> elements
        # Each has a ClassID/CategoryID that identifies the rule
        for issue in root.iter():
            tag_name = issue.tag.replace(ns, "") if ns else issue.tag
            if tag_name not in ("Issue", "IssueInstance"):
                continue

            # Extract rule identifier
            rule_id = (
                issue.get("classId")
                or issue.get("categoryId")
                or issue.get("instanceId")
                or ""
            )
            if not rule_id or rule_id in seen_rules:
                continue
            seen_rules.add(rule_id)

            full_rule_id = f"fortify:{rule_id}"

            # Extract category/severity from child elements
            category_name = ""
            severity_raw = "medium"
            for child in issue:
                child_tag = child.tag.replace(ns, "") if ns else child.tag
                if child_tag in ("Category", "ClassID"):
                    category_name = child.text or ""
                elif child_tag in ("Impact", "Severity", "Friority"):
                    severity_raw = child.text or "medium"

            severity = SEVERITY_MAP.get(str(severity_raw).lower(), "medium")
            category = CATEGORY_MAP.get(
                category_name.lower().strip(), "security"
            )

            tags = ["fortify", "sast", category]
            if category_name:
                tags.append(category_name.lower().replace(" ", "-"))

            self.upsert(
                rule_id=full_rule_id,
                title=f"Fortify {rule_id}: {category_name}"[:500],
                description=f"Fortify rule {rule_id}. Category: {category_name}.",
                severity=severity,
                category=category,
                language="",
                cwe_ids=[],
                tags=tags,
                source_file=fpath,
                rule_content="",
                rule_format="xml",
                metadata={
                    "checker_id": rule_id,
                    "fortify_category": category_name,
                    "impact": severity_raw,
                    "source": "fpr-audit-import",
                },
            )
            count += 1

        return count

    def _parse_csv(self, fpath):
        """Parse a CSV export of Fortify checker definitions."""
        count = 0

        try:
            with open(fpath, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    checker_id = (
                        row.get("CategoryID")
                        or row.get("RuleID")
                        or row.get("Checker")
                        or row.get("checker")
                        or row.get("rule_id")
                        or ""
                    )
                    if not checker_id:
                        continue

                    full_rule_id = f"fortify:{checker_id}"

                    title = (
                        row.get("Category")
                        or row.get("RuleName")
                        or row.get("CheckerName")
                        or row.get("Description")
                        or checker_id
                    )

                    description = (
                        row.get("Description")
                        or row.get("Detail")
                        or row.get("Abstract")
                        or ""
                    )

                    severity_raw = (
                        row.get("Impact")
                        or row.get("Severity")
                        or row.get("Friority")
                        or row.get("Risk")
                        or "medium"
                    )
                    severity = SEVERITY_MAP.get(str(severity_raw).lower(), "medium")

                    cwe_ids = []
                    cwe_raw = row.get("CWE") or row.get("CWEId") or ""
                    if cwe_raw:
                        cwe_nums = re.findall(r"(\d+)", str(cwe_raw))
                        cwe_ids = [f"CWE-{n}" for n in cwe_nums]

                    fortify_category = row.get("Category") or ""
                    category = CATEGORY_MAP.get(
                        fortify_category.lower().strip(), "security"
                    )
                    language = row.get("Language") or row.get("Analyzer") or ""

                    tags = ["fortify", "sast", category]
                    if fortify_category:
                        tags.append(fortify_category.lower().replace(" ", "-"))

                    self.upsert(
                        rule_id=full_rule_id,
                        title=f"Fortify {checker_id}: {title}"[:500],
                        description=description[:2000],
                        severity=severity,
                        category=category,
                        language=language,
                        cwe_ids=cwe_ids,
                        tags=tags,
                        source_file=fpath,
                        rule_content=json.dumps(row, default=str)[:50000],
                        rule_format="csv",
                        metadata={
                            "checker_id": checker_id,
                            "fortify_category": fortify_category,
                            "impact": severity_raw,
                            "source": "csv-import",
                        },
                    )
                    count += 1
        except Exception as e:
            logger.warning(f"[fortify] Failed to parse CSV {fpath}: {e}")

        return count

    def _parse_xlsx(self, fpath):
        """Parse an XLSX export of Fortify checker definitions.

        Requires openpyxl. If not installed, logs a warning and skips.
        """
        try:
            import openpyxl
        except ImportError:
            logger.warning(
                "[fortify] openpyxl not installed — skipping XLSX file. "
                "Install with: pip install openpyxl"
            )
            return 0

        count = 0
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                headers = [str(h or "").strip().lower() for h in rows[0]]

                for row in rows[1:]:
                    row_dict = {
                        headers[i]: (str(row[i]) if row[i] is not None else "")
                        for i in range(min(len(headers), len(row)))
                    }

                    checker_id = (
                        row_dict.get("categoryid")
                        or row_dict.get("ruleid")
                        or row_dict.get("checker")
                        or ""
                    )
                    if not checker_id:
                        continue

                    full_rule_id = f"fortify:{checker_id}"

                    title = (
                        row_dict.get("category")
                        or row_dict.get("rulename")
                        or row_dict.get("description")
                        or checker_id
                    )

                    description = (
                        row_dict.get("description")
                        or row_dict.get("detail")
                        or row_dict.get("abstract")
                        or ""
                    )

                    severity_raw = (
                        row_dict.get("impact")
                        or row_dict.get("severity")
                        or row_dict.get("friority")
                        or "medium"
                    )
                    severity = SEVERITY_MAP.get(str(severity_raw).lower(), "medium")

                    cwe_ids = []
                    cwe_raw = row_dict.get("cwe") or ""
                    if cwe_raw:
                        cwe_nums = re.findall(r"(\d+)", str(cwe_raw))
                        cwe_ids = [f"CWE-{n}" for n in cwe_nums]

                    fortify_category = row_dict.get("category") or ""
                    category = CATEGORY_MAP.get(
                        fortify_category.lower().strip(), "security"
                    )

                    tags = ["fortify", "sast", category]
                    if fortify_category:
                        tags.append(fortify_category.lower().replace(" ", "-"))

                    self.upsert(
                        rule_id=full_rule_id,
                        title=f"Fortify {checker_id}: {title}"[:500],
                        description=description[:2000],
                        severity=severity,
                        category=category,
                        language=row_dict.get("language") or "",
                        cwe_ids=cwe_ids,
                        tags=tags,
                        source_file=fpath,
                        rule_content=json.dumps(row_dict, default=str)[:50000],
                        rule_format="xlsx",
                        metadata={
                            "checker_id": checker_id,
                            "fortify_category": fortify_category,
                            "impact": severity_raw,
                            "source": "xlsx-import",
                        },
                    )
                    count += 1

            wb.close()
        except Exception as e:
            logger.warning(f"[fortify] Failed to parse XLSX {fpath}: {e}")

        return count

    @staticmethod
    def _xml_text(element, tag):
        """Safely extract text from a child element."""
        child = element.find(tag)
        if child is not None and child.text:
            return child.text.strip()
        return ""