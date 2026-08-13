"""Collector for Synopsys/Black Duck Coverity static analysis rules.

Coverity is a commercial static analysis tool by Synopsys (now Black Duck).
The rule set is proprietary and not publicly available in a Git repository.
However, Coverity exports findings in SARIF, JSON, and CSV formats, and
some rule documentation is publicly available at:
  - https://documentation.blackduck.com/bundle/coverity-docs
  - https://documentation.blackduck.com/en-us/coverity.html

This collector uses source_type='file' — it does NOT clone a repository.
Instead, it imports Coverity rules from local/exported files in these
supported formats:

  1. SARIF (.sarif) — Coverity can export findings as SARIF, which includes
     rule metadata (ID, severity, CWE, description).

  2. JSON export — Coverity's cov-format-errors --json output includes
     checker names, categories, and CWE mappings.

  3. XLSX/CSV — Coverity's export to spreadsheet format contains checker
     IDs and descriptions.

To use this collector:
  a. Export rules from Coverity in SARIF or JSON format.
  b. Place the files in the directory specified by the
     COVERITY_RULES_DIR environment variable (default:
     /tmp/cybersagacity-rules/coverity/).
  c. Run the collector — it will parse all .sarif, .json, .csv, and .xlsx
     files in that directory.

The collector can be extended with a proper importer when Synopsys/Black
Duck publishes a public rules repository.
"""

import os
import json
import csv
import logging

from .base import BaseCollector

logger = logging.getLogger(__name__)

# Coverity impact → normalized severity
SEVERITY_MAP = {
    "critical": "critical",
    "high": "high",
    "medium": "medium",
    "moderate": "medium",
    "low": "low",
    "info": "info",
    "informational": "info",
}


class CoverityCollector(BaseCollector):
    name = "coverity"
    display_name = "Coverity (Black Duck)"
    source_type = "file"
    source_url = (
        "https://documentation.blackduck.com/en-us/coverity.html"
    )
    description = (
        "Synopsys/Black Duck Coverity is a commercial static analysis tool "
        "that detects security vulnerabilities, code quality issues, and "
        "concurrency defects in C/C++, Java, C#, JavaScript, Python, and "
        "other languages. Rules (checkers) map to CWEs and include dataflow "
        "analysis, taint tracking, and interprocedural analysis. This "
        "collector imports rules from exported SARIF/JSON/CSV/XLSX files."
    )
    logo_url = "https://www.blackduck.com/favicon.ico"

    def collect_rules(self):
        """Import Coverity rules from local file exports.

        This collector reads from the directory specified by the
        COVERITY_RULES_DIR environment variable (or the default clone_dir).
        Supported file formats: .sarif, .json, .csv, .xlsx

        See the module docstring for instructions on exporting Coverity rules.
        """
        rules_dir = os.environ.get(
            "COVERITY_RULES_DIR",
            self.clone_dir,
        )

        if not os.path.isdir(rules_dir):
            logger.info(
                f"[coverity] Rules directory {rules_dir} not found. "
                f"Export Coverity rules (SARIF/JSON/CSV) and place them there. "
                f"Set COVERITY_RULES_DIR to override."
            )
            return

        count = 0
        for fname in sorted(os.listdir(rules_dir)):
            fpath = os.path.join(rules_dir, fname)
            if not os.path.isfile(fpath):
                continue

            if fname.endswith(".sarif") or (
                fname.endswith(".json") and "coverity" in fname.lower()
            ):
                count += self._parse_sarif(fpath)
            elif fname.endswith(".csv"):
                count += self._parse_csv(fpath)
            elif fname.endswith(".xlsx"):
                count += self._parse_xlsx(fpath)

        logger.info(f"[coverity] Imported {count} rules from {rules_dir}")

    def _parse_sarif(self, fpath):
        """Parse a SARIF file for Coverity rule definitions."""
        count = 0

        try:
            with open(fpath, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError):
            logger.warning(f"[coverity] Failed to parse SARIF: {fpath}")
            return 0

        rel_path = fpath  # file-based, no clone_dir

        # SARIF structure: runs[].tool.driver.rules[]
        for run in data.get("runs", []):
            driver = run.get("tool", {}).get("driver", {})
            rules = driver.get("rules", [])

            for rule in rules:
                rule_id = rule.get("id", "")
                if not rule_id:
                    continue

                # Prefix with coverity namespace
                if not rule_id.startswith("coverity:"):
                    full_rule_id = f"coverity:{rule_id}"
                else:
                    full_rule_id = rule_id

                # Extract metadata
                properties = rule.get("properties", {})
                full_desc = rule.get("fullDescription", {}).get("text", "")
                short_desc = rule.get("shortDescription", {}).get("text", "")
                name = rule.get("name", rule_id)

                # Severity from properties or default level
                level = rule.get("defaultConfiguration", {}).get("level", "warning")
                severity_raw = properties.get("severity", properties.get("impact", level))
                severity = SEVERITY_MAP.get(str(severity_raw).lower(), "medium")

                # CWE from properties
                cwe_ids = []
                cwe_raw = properties.get("cwe", properties.get("cweId", ""))
                if cwe_raw:
                    if isinstance(cwe_raw, list):
                        cwe_ids = [str(c) for c in cwe_raw]
                    else:
                        import re
                        cwe_nums = re.findall(r"(\d+)", str(cwe_raw))
                        cwe_ids = [f"CWE-{n}" for n in cwe_nums]

                # Category
                category = properties.get("category", "static-analysis")
                tags = ["coverity", "sast", category]
                if properties.get("kind"):
                    tags.append(properties["kind"])

                # OWASP
                owasp_ids = []
                owasp_raw = properties.get("owasp", "")
                if owasp_raw:
                    if isinstance(owasp_raw, list):
                        owasp_ids = [str(o) for o in owasp_raw]
                    else:
                        owasp_ids = [str(owasp_raw)]

                self.upsert(
                    rule_id=full_rule_id,
                    title=f"Coverity {name}: {short_desc}"[:500],
                    description=full_desc or short_desc,
                    severity=severity,
                    category=category,
                    language=properties.get("language", ""),
                    cwe_ids=cwe_ids,
                    owasp_ids=owasp_ids,
                    tags=tags,
                    source_file=rel_path,
                    rule_content=json.dumps(rule, indent=2, default=str)[:50000],
                    rule_format="sarif",
                    metadata={
                        "checker_id": rule_id,
                        "checker_name": name,
                        "kind": properties.get("kind", ""),
                        "impact": properties.get("impact", ""),
                        "subcategory": properties.get("subcategory", ""),
                        "source": "sarif-import",
                    },
                )
                count += 1

        return count

    def _parse_csv(self, fpath):
        """Parse a CSV export of Coverity checker definitions."""
        count = 0

        try:
            with open(fpath, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Common CSV columns from Coverity exports
                    checker_id = (
                        row.get("Checker")
                        or row.get("checker")
                        or row.get("CheckerId")
                        or row.get("rule_id")
                        or ""
                    )
                    if not checker_id:
                        continue

                    full_rule_id = f"coverity:{checker_id}"

                    title = (
                        row.get("CheckerName")
                        or row.get("Subcategory")
                        or row.get("Description")
                        or checker_id
                    )

                    description = (
                        row.get("Description")
                        or row.get("SubcategoryLong")
                        or row.get("Detail")
                        or ""
                    )

                    severity_raw = (
                        row.get("Impact")
                        or row.get("Severity")
                        or row.get("Risk")
                        or "medium"
                    )
                    severity = SEVERITY_MAP.get(str(severity_raw).lower(), "medium")

                    # CWE
                    cwe_ids = []
                    cwe_raw = row.get("CWE") or row.get("CWEId") or ""
                    if cwe_raw:
                        import re
                        cwe_nums = re.findall(r"(\d+)", str(cwe_raw))
                        cwe_ids = [f"CWE-{n}" for n in cwe_nums]

                    category = row.get("Category") or "static-analysis"
                    language = row.get("Language") or ""

                    self.upsert(
                        rule_id=full_rule_id,
                        title=f"Coverity {checker_id}: {title}"[:500],
                        description=description[:2000],
                        severity=severity,
                        category=category,
                        language=language,
                        cwe_ids=cwe_ids,
                        tags=["coverity", "sast", category],
                        source_file=fpath,
                        rule_content=json.dumps(row, default=str)[:50000],
                        rule_format="csv",
                        metadata={
                            "checker_id": checker_id,
                            "impact": severity_raw,
                            "source": "csv-import",
                        },
                    )
                    count += 1
        except Exception as e:
            logger.warning(f"[coverity] Failed to parse CSV {fpath}: {e}")

        return count

    def _parse_xlsx(self, fpath):
        """Parse an XLSX export of Coverity checker definitions.

        Requires openpyxl. If not installed, logs a warning and skips.
        """
        try:
            import openpyxl
        except ImportError:
            logger.warning(
                "[coverity] openpyxl not installed — skipping XLSX file. "
                "Install with: pip install openpyxl"
            )
            return 0

        count = 0
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                # First row = headers
                headers = [str(h or "").strip().lower() for h in rows[0]]

                for row in rows[1:]:
                    row_dict = {
                        headers[i]: (str(row[i]) if row[i] is not None else "")
                        for i in range(min(len(headers), len(row)))
                    }

                    checker_id = (
                        row_dict.get("checker")
                        or row_dict.get("checkerid")
                        or row_dict.get("rule_id")
                        or ""
                    )
                    if not checker_id:
                        continue

                    full_rule_id = f"coverity:{checker_id}"

                    title = (
                        row_dict.get("checkername")
                        or row_dict.get("subcategory")
                        or row_dict.get("description")
                        or checker_id
                    )

                    description = (
                        row_dict.get("description")
                        or row_dict.get("subcategorylong")
                        or row_dict.get("detail")
                        or ""
                    )

                    severity_raw = (
                        row_dict.get("impact")
                        or row_dict.get("severity")
                        or "medium"
                    )
                    severity = SEVERITY_MAP.get(str(severity_raw).lower(), "medium")

                    cwe_ids = []
                    cwe_raw = row_dict.get("cwe") or ""
                    if cwe_raw:
                        import re
                        cwe_nums = re.findall(r"(\d+)", str(cwe_raw))
                        cwe_ids = [f"CWE-{n}" for n in cwe_nums]

                    self.upsert(
                        rule_id=full_rule_id,
                        title=f"Coverity {checker_id}: {title}"[:500],
                        description=description[:2000],
                        severity=severity,
                        category=row_dict.get("category") or "static-analysis",
                        language=row_dict.get("language") or "",
                        cwe_ids=cwe_ids,
                        tags=["coverity", "sast"],
                        source_file=fpath,
                        rule_content=json.dumps(row_dict, default=str)[:50000],
                        rule_format="xlsx",
                        metadata={
                            "checker_id": checker_id,
                            "impact": severity_raw,
                            "source": "xlsx-import",
                        },
                    )
                    count += 1

            wb.close()
        except Exception as e:
            logger.warning(f"[coverity] Failed to parse XLSX {fpath}: {e}")

        return count